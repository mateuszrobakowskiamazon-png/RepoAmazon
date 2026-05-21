"""
Amazon Multi-Language - Generator Tytułów (Streamlit)
Obsługuje DE / FR / IT / ES / NL / SE / EN / PL.
Dodatkowo: tłumaczenie wygenerowanego tytułu na inny język (słownik strukturalny).
"""

import re
from collections import Counter
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from languages import CATEGORIES, LANGUAGES, TRANSLATIONS


# ===================== STAŁE =====================

FORBIDDEN_CHARS = "!$?_{}^¬¦|"
TITLE_LIMIT = 200


# ===================== PARSOWANIE =====================

def parse_cerebro(uploaded_file) -> pd.DataFrame:
    df = pd.read_excel(uploaded_file)
    expected = ["Keyword Phrase", "Search Volume", "Keyword Sales"]
    if not all(c in df.columns for c in expected):
        raise ValueError(
            f"Plik nie wygląda na export z Cerebro - brak kolumn: {expected}"
        )
    return (
        df[expected + [c for c in df.columns if c.startswith("B0")]]
        .sort_values("Search Volume", ascending=False)
        .reset_index(drop=True)
    )


def parse_competitor_titles(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line.strip()]


def extract_keywords_from_titles(titles: list[str], exempt: set) -> list[tuple[str, int]]:
    all_words = []
    for t in titles:
        words = re.findall(r"[A-Za-zÄÖÜäöüßéèêàâçñíóúÁÉÍÓÚłąęśćźżńŁŚŻŹĆĄĘŃÅÄÖå]+", t.lower())
        all_words.extend(w for w in words if w not in exempt and len(w) > 3)
    return Counter(all_words).most_common(20)


# ===================== AUTO-EKSTRAKCJA Z KONKURENCJI + CEREBRO =====================

def build_word_to_de_key_map() -> dict:
    """Mapa: dowolne słowo (w dowolnym języku, lowercase) → klucz niemiecki."""
    mapping = {}
    for de_word, langs in TRANSLATIONS.items():
        mapping[de_word.lower()] = de_word
        for _, translation in langs.items():
            mapping[translation.lower()] = de_word
            # Dla wielowyrazowych ("Coussin de Chaise") - dodaj też pierwsze słowo
            first_word = translation.split()[0].lower()
            if first_word not in mapping:
                mapping[first_word] = de_word
    return mapping


def classify_word(word: str, word_map: dict) -> tuple[str | None, str | None]:
    """Zwraca (kategoria, klucz_DE) lub (None, None) jeśli nie rozpoznane."""
    word_lower = word.lower().strip(".,;:")
    if word_lower in word_map:
        de_key = word_map[word_lower]
        return CATEGORIES.get(de_key), de_key
    return None, None


def extract_from_competitor_titles(
    titles: list[str], lang: str, main_kw_words: set
) -> dict:
    """
    Analizuje tytuły konkurencji, wyciąga słowa per kategoria.
    Zwraca dict {synonyms, contexts, features, materials, colors, shape}.
    Słowa wyrażone już w main_kw są pomijane (nie chcemy duplikatów).
    """
    word_map = build_word_to_de_key_map()

    categorized = {
        "products": Counter(),   # → synonimy
        "context": Counter(),
        "feature": Counter(),
        "material": Counter(),
        "color": Counter(),
        "shape": Counter(),
    }

    # Tokenizuj wszystkie tytuły, klasyfikuj słowa
    for title in titles:
        # Dla 2-wyrazowych fraz typu "Memory Foam", "Cuscino Sedia" - sprawdź również pary
        words = re.findall(r"[A-Za-zÄÖÜäöüßéèêàâçñíóúÁÉÍÓÚłąęśćźżńŁŚŻŹĆĄĘŃÅÄÖå]+", title)
        seen_in_title = set()

        # Najpierw bigramy
        for i in range(len(words) - 1):
            bigram = f"{words[i]} {words[i+1]}".lower()
            if bigram in word_map:
                de_key = word_map[bigram]
                cat = CATEGORIES.get(de_key)
                if cat and de_key not in seen_in_title:
                    seen_in_title.add(de_key)
                    target_cat = "products" if cat == "product" else cat
                    if target_cat in categorized:
                        # Tłumaczymy klucz DE na język docelowy
                        translated = TRANSLATIONS.get(de_key, {}).get(lang, de_key) if lang != "DE" else de_key
                        categorized[target_cat][translated] += 1

        # Potem pojedyncze słowa
        for word in words:
            cat, de_key = classify_word(word, word_map)
            if cat and de_key not in seen_in_title:
                seen_in_title.add(de_key)
                target_cat = "products" if cat == "product" else cat
                if target_cat in categorized:
                    if de_key.lower() in main_kw_words:
                        continue  # nie duplikuj main_kw
                    translated = TRANSLATIONS.get(de_key, {}).get(lang, de_key) if lang != "DE" else de_key
                    categorized[target_cat][translated] += 1

    # Wybierz słowa które pojawiają się co najmniej raz, posortowane po częstości
    def top_words(counter: Counter, min_count: int = 1, max_n: int = 8) -> list[str]:
        return [w for w, c in counter.most_common(max_n) if c >= min_count]

    return {
        "synonyms": top_words(categorized["products"], min_count=1, max_n=6),
        "contexts": top_words(categorized["context"], min_count=1, max_n=6),
        "features": top_words(categorized["feature"], min_count=1, max_n=4),
        "materials": top_words(categorized["material"], min_count=1, max_n=2),
        "colors": top_words(categorized["color"], min_count=1, max_n=2),
        "shapes": top_words(categorized["shape"], min_count=1, max_n=1),
    }


def extract_main_kw_from_cerebro(df: pd.DataFrame, top_n: int = 3) -> list[str]:
    """Wyciąga top N keywordów po Search Volume, w Title Case."""
    if df.empty:
        return []
    return [kw.title() for kw in df.head(top_n)["Keyword Phrase"].tolist()]


def extract_dimensions(text: str) -> str | None:
    """Wyciąga wymiar typu '40 cm', '120x80', 'ø35cm' z tekstu."""
    # 40 cm, 40cm
    m = re.search(r"\b(\d{2,3})\s*cm\b", text, re.IGNORECASE)
    if m:
        return f"{m.group(1)} cm"
    # 120x80, 120 x 80
    m = re.search(r"\b(\d{2,3})\s*x\s*(\d{2,3})\b", text, re.IGNORECASE)
    if m:
        return f"{m.group(1)}x{m.group(2)}"
    # ø35, ø 35
    m = re.search(r"ø\s*(\d{2,3})", text, re.IGNORECASE)
    if m:
        return f"{m.group(1)} cm"
    return None


def extract_set_size(text: str, lang: str) -> str | None:
    """Wyciąga 'set/lot/pack' z tytułu."""
    patterns = [
        (r"\b(\d+)er\s*Set\b", "DE"),
        (r"\bLot\s*de\s*(\d+)\b", "FR"),
        (r"\bSet\s*da\s*(\d+)\b", "IT"),
        (r"\bPack\s*de\s*(\d+)\b", "ES"),
        (r"\bSet\s*van\s*(\d+)\b", "NL"),
        (r"\bSet\s*om\s*(\d+)\b", "SE"),
        (r"\bSet\s*of\s*(\d+)\b", "EN"),
        (r"\bZestaw\s*(\d+)\s*szt\b", "PL"),
        (r"\b(\d+)er-Set\b", "DE"),
        (r"\b(\d+)-pack\b", "EN"),
    ]
    for pat, _ in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            num = m.group(1)
            key = f"{num} szt."
            if key in LANGUAGES[lang]["set_sizes"]:
                return LANGUAGES[lang]["set_sizes"][key]
    return None


def auto_extract_all(
    cerebro_df: pd.DataFrame,
    competitor_titles: list[str],
    lang: str,
    user_main_kw: str = "",
) -> dict:
    """
    Pełna ekstrakcja - łączy dane z Cerebro i tytułów konkurencji.
    Zwraca słownik z polami do auto-wypełnienia w UI.
    """
    result = {
        "main_kw_suggestions": [],
        "main_kw": "",
        "size": "",
        "set_size": "",
        "material": "",
        "color": "",
        "shape": "",
        "synonyms": [],
        "contexts": [],
        "features": [],
    }

    # Sugestie main_kw z Cerebro (top SV)
    cerebro_top = extract_main_kw_from_cerebro(cerebro_df, top_n=3)
    result["main_kw_suggestions"] = cerebro_top
    if cerebro_top and not user_main_kw:
        result["main_kw"] = cerebro_top[0]

    main_kw_to_use = user_main_kw or result["main_kw"]
    main_kw_words = set(main_kw_to_use.lower().split())

    # Analiza tytułów konkurencji
    if competitor_titles:
        extracted = extract_from_competitor_titles(competitor_titles, lang, main_kw_words)
        result["synonyms"] = extracted["synonyms"]
        result["contexts"] = extracted["contexts"]
        result["features"] = extracted["features"]
        result["material"] = extracted["materials"][0] if extracted["materials"] else ""
        result["color"] = extracted["colors"][0] if extracted["colors"] else ""
        result["shape"] = extracted["shapes"][0] if extracted["shapes"] else ""

        # Wymiar i set - z najczęstszego tytułu (pierwszego)
        for title in competitor_titles:
            if not result["size"]:
                size = extract_dimensions(title)
                if size:
                    result["size"] = size
            if not result["set_size"]:
                ss = extract_set_size(title, lang)
                if ss:
                    result["set_size"] = ss
            if result["size"] and result["set_size"]:
                break

    return result


# ===================== WALIDACJA =====================

def validate_title(title: str, lang: str, no_commas: bool = True) -> dict:
    issues = []
    if len(title) > TITLE_LIMIT:
        issues.append(f"Za długi: {len(title)}/{TITLE_LIMIT} znaków")

    bad = [c for c in FORBIDDEN_CHARS if c in title]
    if bad:
        issues.append(f"Zakazane znaki: {' '.join(bad)}")

    if no_commas and "," in title:
        issues.append("Zawiera przecinki (wyłączone w ustawieniach)")

    exempt = LANGUAGES[lang]["exempt_words"]
    words = re.findall(r"[A-Za-zÄÖÜäöüßéèêàâçñíóúÁÉÍÓÚłąęśćźżńŁŚŻŹĆĄĘŃÅÄÖå]+", title.lower())
    counts = Counter(w for w in words if w not in exempt and len(w) > 1)
    repeated = {w: c for w, c in counts.items() if c > 2}
    if repeated:
        issues.append(f"Powtórzenia 3+ razy: {repeated}")

    return {"length": len(title), "valid": not issues, "issues": issues}


# ===================== GENERATORY TYTUŁÓW =====================

def build_title(
    style: str,
    lang: str,
    brand: str,
    main_kw: str,
    size: str,
    set_size: str,
    material: str,
    features: list[str],
    synonyms: list[str],
    contexts: list[str],
    color: str,
    no_commas: bool,
) -> str:
    """style: 'A' (zlepek), 'B' (hybryda), 'C' (zdaniowa)"""
    cfg = LANGUAGES[lang]
    conn = cfg["connectors"]
    dash = " – "
    amp = " & "
    feat_adj = cfg.get("feature_adj", {})

    syns = list(dict.fromkeys(s.strip() for s in synonyms if s.strip()))
    ctxs = list(dict.fromkeys(c.strip() for c in contexts if c.strip()))
    feats = [f.strip() for f in features if f.strip()]

    if style == "A":
        parts = [brand, main_kw, size, set_size, material, dash, *syns, *feats]
        if ctxs:
            parts.append(conn["for"])
            parts.extend(ctxs)
        title = " ".join(p for p in parts if p)
        if color:
            title += dash + color
        return title.replace("  ", " ").strip()

    if style == "B":
        feat_str = " ".join(feats)
        if len(syns) > 2:
            syn_block = amp.join(syns[:2]) + " " + " ".join(syns[2:])
        elif syns:
            syn_block = amp.join(syns)
        else:
            syn_block = ""

        if len(ctxs) > 1:
            ctx_block = " ".join(ctxs[:-1]) + amp + ctxs[-1]
        elif ctxs:
            ctx_block = ctxs[0]
        else:
            ctx_block = ""

        parts = [brand, main_kw, size, set_size]
        if material:
            parts.extend([conn["from"], material])
        if feat_str:
            parts.append(feat_str)
        if syn_block:
            parts.extend([dash, syn_block])
        if ctx_block:
            parts.extend([conn["for"], ctx_block])
        title = " ".join(p for p in parts if p)
        if color:
            title += dash + color
        return title.replace("  ", " ").strip()

    if style == "C":
        if feats:
            first_feat = feats[0].lower()
            adj_form = feat_adj.get(first_feat, feats[0])
        else:
            adj_form = ""
        main_syn = syns[0] if syns else main_kw
        other_syns = syns[1:]
        if len(ctxs) > 1:
            ctx_str = " ".join(ctxs[:-1]) + f" {conn['and']} " + ctxs[-1]
        elif ctxs:
            ctx_str = ctxs[0]
        else:
            ctx_str = ""

        parts = [brand, main_kw, size]
        if material:
            parts.extend([conn["from"], material])
        if set_size:
            parts.append(set_size)
        parts.append(dash)
        if adj_form:
            parts.append(adj_form)
        parts.append(main_syn)
        if ctx_str:
            parts.extend([conn["for"], ctx_str])
        if other_syns:
            parts.extend([dash, amp.join(other_syns[:2])])

        title = " ".join(p for p in parts if p)
        if color:
            title += dash + color
        return title.replace("  ", " ").strip()

    raise ValueError(f"Nieznany styl: {style}")


def trim_to_fit(style: str, max_len: int = TITLE_LIMIT, **kwargs) -> str:
    """Auto-przycinanie kontekstów i synonimów żeby zmieścić w 200 zn."""
    title = build_title(style=style, **kwargs)
    if len(title) <= max_len:
        return title

    contexts = list(kwargs.get("contexts", []))
    while contexts and len(title) > max_len:
        contexts.pop()
        kwargs["contexts"] = contexts
        title = build_title(style=style, **kwargs)

    syns = list(kwargs.get("synonyms", []))
    while len(syns) > 2 and len(title) > max_len:
        syns.pop()
        kwargs["synonyms"] = syns
        title = build_title(style=style, **kwargs)

    return title


# ===================== TŁUMACZENIE TYTUŁU =====================

def build_reverse_translations() -> dict:
    """Buduje słownik wieloznaczny: dowolne słowo (w dowolnym języku) → klucz DE."""
    reverse = {}
    for de_word, langs in TRANSLATIONS.items():
        reverse[de_word.lower()] = de_word
        for _, trans in langs.items():
            reverse[trans.lower()] = de_word
    return reverse


def translate_title(title: str, from_lang: str, to_lang: str) -> tuple[str, list[str]]:
    """
    Tłumaczy tytuł na docelowy język.
    Strategia: identyfikuje znane słowa (przez słownik DE-jako-pivot), mapuje
    łączniki/set/separatory, pozostałe słowa zostawia z oznaczeniem.
    Zwraca: (przetłumaczony tytuł, lista nieznanych słów).
    """
    if from_lang == to_lang:
        return title, []

    src_cfg = LANGUAGES[from_lang]
    dst_cfg = LANGUAGES[to_lang]

    # Mapowanie set_sizes: szukamy fragmentu set_size source i zamieniamy na target
    for key, src_val in src_cfg["set_sizes"].items():
        dst_val = dst_cfg["set_sizes"][key]
        if src_val in title:
            title = title.replace(src_val, dst_val)
            break

    # Mapowanie łączników (für → pour itp.)
    src_conn = src_cfg["connectors"]
    dst_conn = dst_cfg["connectors"]
    for key in ["for", "from", "and", "or"]:
        src_w = src_conn[key]
        dst_w = dst_conn[key]
        # Zamieniamy całe słowa (z ograniczeniami granic słowa)
        title = re.sub(rf"\b{re.escape(src_w)}\b", dst_w, title, flags=re.IGNORECASE)

    # Słownikowe tłumaczenie - słowo po słowie z zachowaniem separatorów
    reverse_dict = build_reverse_translations()
    unknown_words = []

    def translate_word(match):
        word = match.group(0)
        word_lower = word.lower()
        if word_lower in reverse_dict:
            de_key = reverse_dict[word_lower]
            if to_lang == "DE":
                return de_key
            translations = TRANSLATIONS.get(de_key, {})
            if to_lang in translations:
                return translations[to_lang]
        # Nie znamy - pomijamy z listy nieznanych jeśli to nie jest exempt
        if (word_lower not in src_cfg["exempt_words"]
                and word_lower not in dst_cfg["exempt_words"]
                and len(word) > 2
                and not word_lower.isdigit()
                and not re.match(r"^\d+x\d+$", word_lower)
                and word_lower not in {"cm", "mm", "kg", "g", "ml"}):
            unknown_words.append(word)
        return word

    # Tokenizuj po słowach (zachowując spacje i znaki interpunkcyjne)
    pattern = r"[A-Za-zÄÖÜäöüßéèêàâçñíóúÁÉÍÓÚłąęśćźżńŁŚŻŹĆĄĘŃÅÄÖå]+"
    translated = re.sub(pattern, translate_word, title)

    # Usuń duplikaty z nieznanych
    unknown_words = list(dict.fromkeys(unknown_words))

    return translated.replace("  ", " ").strip(), unknown_words


# ===================== EXCEL OUTPUT =====================

def build_excel(
    titles: dict[str, dict],
    cerebro_df: pd.DataFrame,
    competitor_titles: list[str],
    backend_terms: str,
    product_params: dict,
    translated: dict | None = None,
) -> bytes:
    wb = Workbook()
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # === Arkusz 1: Opcje tytułów ===
    ws = wb.active
    ws.title = "Opcje tytułów"
    headers = ["Opcja", "Tytuł", "Długość", "Walidacja", "Problemy"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial")
        c.fill = PatternFill("solid", start_color="2C3E50")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    style_desc = {
        "A": "A – Klasyczny zlepek",
        "B": "B – Hybryda (REKOMENDOWANE)",
        "C": "C – Zdaniowa naturalna",
    }
    fills = {"A": "FFF3CD", "B": "D4EDDA", "C": "D1ECF1"}

    for r_idx, style in enumerate(["A", "B", "C"], 2):
        data = titles[style]
        ws.cell(row=r_idx, column=1, value=style_desc[style])
        ws.cell(row=r_idx, column=2, value=data["title"])
        ws.cell(row=r_idx, column=3, value=f"{data['length']} zn.")
        ws.cell(row=r_idx, column=4, value="✅ OK" if data["valid"] else "❌ Problem")
        ws.cell(row=r_idx, column=5, value="\n".join(data["issues"]) if data["issues"] else "")
        for col in range(1, 6):
            cell = ws.cell(row=r_idx, column=col)
            cell.font = Font(name="Arial", size=10, bold=(col == 1))
            cell.fill = PatternFill("solid", start_color=fills[style])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 35
    for r in range(2, 5):
        ws.row_dimensions[r].height = 90

    # === Arkusz 2: Tłumaczenia (jeśli są) ===
    if translated:
        ws_t = wb.create_sheet("Tłumaczenia")
        ws_t["A1"] = "Tytuły przetłumaczone na inne języki"
        ws_t["A1"].font = Font(bold=True, size=12, name="Arial")
        ws_t.cell(row=3, column=1, value="Język").font = Font(bold=True, name="Arial")
        ws_t.cell(row=3, column=2, value="Opcja").font = Font(bold=True, name="Arial")
        ws_t.cell(row=3, column=3, value="Tytuł").font = Font(bold=True, name="Arial")
        ws_t.cell(row=3, column=4, value="Długość").font = Font(bold=True, name="Arial")
        ws_t.cell(row=3, column=5, value="Nieznane słowa").font = Font(bold=True, name="Arial")

        row = 4
        for lang_code, data in translated.items():
            for style, info in data["titles"].items():
                ws_t.cell(row=row, column=1, value=LANGUAGES[lang_code]["name"])
                ws_t.cell(row=row, column=2, value=style)
                ws_t.cell(row=row, column=3, value=info["title"])
                ws_t.cell(row=row, column=4, value=f"{len(info['title'])} zn.")
                ws_t.cell(row=row, column=5, value=", ".join(info.get("unknown", [])))
                for col in range(1, 6):
                    ws_t.cell(row=row, column=col).font = Font(name="Arial", size=10)
                    ws_t.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
                    ws_t.cell(row=row, column=col).border = border
                row += 1
        ws_t.column_dimensions["A"].width = 25
        ws_t.column_dimensions["B"].width = 10
        ws_t.column_dimensions["C"].width = 70
        ws_t.column_dimensions["D"].width = 12
        ws_t.column_dimensions["E"].width = 35

    # === Arkusz 3: Top słowa kluczowe (jeśli Cerebro wgrane) ===
    if not cerebro_df.empty:
        ws2 = wb.create_sheet("Top słowa kluczowe")
        for col, h in enumerate(["#", "Słowo kluczowe", "Search Volume", "Keyword Sales"], 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial")
            c.fill = PatternFill("solid", start_color="2C3E50")
            c.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, (_, row_data) in enumerate(cerebro_df.head(20).iterrows(), 2):
            ws2.cell(row=r_idx, column=1, value=r_idx - 1)
            ws2.cell(row=r_idx, column=2, value=row_data["Keyword Phrase"])
            ws2.cell(row=r_idx, column=3, value=int(row_data["Search Volume"]) if pd.notna(row_data["Search Volume"]) else "-")
            ws2.cell(row=r_idx, column=4, value=int(row_data["Keyword Sales"]) if pd.notna(row_data["Keyword Sales"]) else "-")
            for col in range(1, 5):
                cell = ws2.cell(row=r_idx, column=col)
                cell.font = Font(name="Arial", size=10)
                cell.border = border
                if r_idx <= 4:
                    cell.fill = PatternFill("solid", start_color="FFEAA7")

        ws2.column_dimensions["A"].width = 5
        ws2.column_dimensions["B"].width = 40
        ws2.column_dimensions["C"].width = 15
        ws2.column_dimensions["D"].width = 15

    # === Arkusz 4: Backend ===
    if backend_terms:
        ws3 = wb.create_sheet("Backend Search Terms")
        ws3["A1"] = "Backend Search Terms (Seller Central → Schlüsselwörter / Keywords)"
        ws3["A1"].font = Font(bold=True, size=12, name="Arial")
        ws3["A2"] = "Wklej do pola „Allgemeine Schlüsselwörter / General Keywords”. Limit Amazon: 249 znaków."
        ws3["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
        ws3["A2"].alignment = Alignment(wrap_text=True)
        ws3["A4"] = backend_terms
        ws3["A4"].font = Font(name="Arial", size=10)
        ws3["A4"].alignment = Alignment(wrap_text=True, vertical="top")
        ws3["A4"].fill = PatternFill("solid", start_color="F0F0F0")
        ws3["A5"] = f"Długość: {len(backend_terms)} znaków / 249"
        ws3["A5"].font = Font(italic=True, size=9, color="666666", name="Arial")
        ws3.column_dimensions["A"].width = 90
        ws3.row_dimensions[4].height = 60

    # === Arkusz 5: Konkurencja ===
    if competitor_titles:
        ws4 = wb.create_sheet("Konkurencja")
        ws4["A1"] = "Tytuły konkurencji (wklejone)"
        ws4["A1"].font = Font(bold=True, size=12, name="Arial")
        for i, t in enumerate(competitor_titles, 3):
            ws4.cell(row=i, column=1, value=f"#{i-2}")
            ws4.cell(row=i, column=2, value=t)
            ws4.cell(row=i, column=3, value=f"{len(t)} zn.")
            for col in range(1, 4):
                ws4.cell(row=i, column=col).font = Font(name="Arial", size=10)
                ws4.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws4.column_dimensions["A"].width = 5
        ws4.column_dimensions["B"].width = 90
        ws4.column_dimensions["C"].width = 12

    # === Arkusz 6: Parametry ===
    ws5 = wb.create_sheet("Parametry produktu")
    ws5["A1"] = "Parametry użyte do generowania tytułów"
    ws5["A1"].font = Font(bold=True, size=12, name="Arial")
    for i, (key, val) in enumerate(product_params.items(), 3):
        ws5.cell(row=i, column=1, value=key).font = Font(name="Arial", size=10, bold=True)
        ws5.cell(row=i, column=2, value=str(val)).font = Font(name="Arial", size=10)
    ws5.column_dimensions["A"].width = 25
    ws5.column_dimensions["B"].width = 70

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ===================== STREAMLIT UI =====================

st.set_page_config(
    page_title="Amazon Multi-Language – Generator Tytułów",
    page_icon="📝",
    layout="wide",
)

st.title("📝 Amazon Multi-Language – Generator Tytułów")
st.caption(
    "Wybierz język rynku, wpisz parametry produktu (większość opcjonalnych) "
    "i wygeneruj 3 wersje tytułu. Plik Cerebro i tytuły konkurencji są opcjonalne. "
    "Po wygenerowaniu możesz przetłumaczyć tytuł na inny język rynku."
)

# Inicjalizacja session state
if "generated" not in st.session_state:
    st.session_state.generated = False
if "titles_data" not in st.session_state:
    st.session_state.titles_data = {}
if "translated_data" not in st.session_state:
    st.session_state.translated_data = {}
if "auto_filled" not in st.session_state:
    st.session_state.auto_filled = {}

# === Sidebar ===
with st.sidebar:
    st.header("🌍 Język / Kraj")
    lang_code = st.selectbox(
        "Język tytułu",
        list(LANGUAGES.keys()),
        format_func=lambda x: LANGUAGES[x]["name"],
        index=0,
    )

    st.markdown("---")
    st.header("🏷️ Parametry produktu")
    st.caption("⭐ = wymagane. Pozostałe wypełnij ręcznie lub użyj „🤖 Auto-wypełnij” na górze.")

    af = st.session_state.auto_filled

    brand = st.text_input("⭐ Marka", placeholder="np. BIELIK", key="brand_input")
    main_kw = st.text_input(
        "⭐ Główna fraza produktu",
        value=af.get("main_kw", ""),
        placeholder="np. Sitzkissen Rund",
        key="main_kw_input",
        help="Z Cerebro: top fraza po Search Volume. Auto-wypełniana po kliknięciu „Auto-wypełnij”.",
    )
    size = st.text_input(
        "Wymiar",
        value=af.get("size", ""),
        placeholder="np. 40 cm albo 120x80",
        key="size_input",
    )

    set_options = list(LANGUAGES[lang_code]["set_sizes"].keys())
    # Wybierz domyślny set na podstawie auto-wypełnienia
    auto_set = af.get("set_size", "")
    default_index = 0
    if auto_set:
        for i, key in enumerate(set_options):
            if LANGUAGES[lang_code]["set_sizes"][key] == auto_set:
                default_index = i + 1  # +1 bo dodajemy "(brak)" na początku
                break

    set_key = st.selectbox(
        "Set / ilość",
        ["(brak)"] + set_options,
        index=default_index,
        key="set_input",
    )
    set_size = "" if set_key == "(brak)" else LANGUAGES[lang_code]["set_sizes"][set_key]

    material = st.text_input(
        "Materiał",
        value=af.get("material", ""),
        placeholder="np. Cord, Filz, Cotton",
        key="material_input",
    )

    features_text = st.text_area(
        "Cechy (jedna per linia)",
        value="\n".join(af.get("features", [])),
        placeholder="waschbar\nrutschfest",
        height=80,
        key="features_input",
    )

    synonyms_text = st.text_area(
        "Synonimy (jeden per linia)",
        value="\n".join(af.get("synonyms", [])),
        placeholder="Stuhlkissen\nBodenkissen",
        height=120,
        key="synonyms_input",
    )

    contexts_text = st.text_area(
        "Konteksty (jeden per linia)",
        value="\n".join(af.get("contexts", [])),
        placeholder="Esszimmer\nKüche\nGarten",
        height=120,
        key="contexts_input",
    )

    color = st.text_input(
        "Kolor",
        value=af.get("color", ""),
        placeholder="np. Anthrazit",
        key="color_input",
    )

    st.markdown("---")
    no_commas = st.toggle("Bez przecinków", value=True)
    is_parent = st.toggle(
        "PARENT-tytuł (wariant kolorystyczny)",
        value=False,
        help="Parent NIE może mieć koloru w tytule",
    )

# === Główne pole ===
st.subheader("📥 Wgraj dane wejściowe")
col1, col2 = st.columns([1, 1])

with col1:
    cerebro_file = st.file_uploader(
        "📁 Plik Cerebro (Helium10) - .xlsx",
        type=["xlsx"],
        help="Eksport z Cerebro z keywordami, Search Volume, Keyword Sales",
    )
    if cerebro_file:
        st.success(f"✅ Wgrano: {cerebro_file.name}")

with col2:
    competitor_text = st.text_area(
        "📋 Tytuły konkurencji (jeden per linia)",
        placeholder="Lsjoaw Sitzkissen Rund 35cm Dicke 6cm 1 Stück Bodenkissen...\nheimtexland Sitzkissen Set Filz Stuhlkissen Type 631...",
        height=150,
        key="competitor_text_input",
    )

# === Przycisk Auto-wypełnij ===
if cerebro_file or competitor_text.strip():
    if st.button(
        "🤖 Auto-wypełnij parametry z Cerebro + konkurencji",
        use_container_width=True,
        type="secondary",
    ):
        # Parsuj Cerebro
        if cerebro_file:
            try:
                cerebro_df = parse_cerebro(cerebro_file)
                st.session_state.cerebro_df = cerebro_df
            except ValueError as e:
                st.error(f"❌ {e}")
                cerebro_df = pd.DataFrame()
        else:
            cerebro_df = pd.DataFrame(columns=["Keyword Phrase", "Search Volume", "Keyword Sales"])
            st.session_state.cerebro_df = cerebro_df

        # Parsuj konkurencję
        competitor_titles = parse_competitor_titles(competitor_text)
        st.session_state.competitor_titles = competitor_titles

        # Auto-ekstrakcja
        existing_main_kw = st.session_state.get("main_kw_input", "")
        extracted = auto_extract_all(
            cerebro_df, competitor_titles, lang_code, existing_main_kw
        )

        st.session_state.auto_filled = extracted

        # Komunikat sukcesu z podsumowaniem
        summary = []
        if extracted["main_kw"]:
            summary.append(f"główna fraza: **{extracted['main_kw']}**")
        if extracted["synonyms"]:
            summary.append(f"{len(extracted['synonyms'])} synonimów")
        if extracted["contexts"]:
            summary.append(f"{len(extracted['contexts'])} kontekstów")
        if extracted["features"]:
            summary.append(f"{len(extracted['features'])} cech")
        if extracted["material"]:
            summary.append(f"materiał: **{extracted['material']}**")

        st.success(
            f"✅ Auto-wypełniono parametry: {', '.join(summary)}. "
            "Sprawdź sidebar po lewej i edytuj jeśli chcesz, potem kliknij **Generuj**."
        )

        # Pokaż top SV z Cerebro
        if extracted["main_kw_suggestions"]:
            with st.expander("📊 Sugestie głównej frazy (top SV z Cerebro)", expanded=True):
                for i, kw in enumerate(extracted["main_kw_suggestions"], 1):
                    st.write(f"{i}. **{kw}**")

        st.rerun()

st.markdown("---")
st.subheader("🔍 Słowa do backendu (opcjonalnie)")
backend_text = st.text_area(
    "Słowa do pola Schlüsselwörter / Keywords (do 249 zn.)",
    placeholder="np. dodatkowe synonimy lub literówki",
    height=70,
    key="backend_input",
)

# === Przycisk Generuj ===
st.markdown("---")
if st.button("🚀 Generuj tytuły", type="primary", use_container_width=True):
    if not brand or not main_kw:
        st.error("❌ Marka i Główna fraza są wymagane.")
        st.stop()

    # Cerebro - z session_state jeśli już sparsowane, lub świeże
    if "cerebro_df" in st.session_state and not st.session_state.cerebro_df.empty:
        cerebro_df = st.session_state.cerebro_df
    elif cerebro_file:
        try:
            cerebro_df = parse_cerebro(cerebro_file)
            st.session_state.cerebro_df = cerebro_df
        except ValueError as e:
            st.error(f"❌ {e}")
            st.stop()
    else:
        cerebro_df = pd.DataFrame(columns=["Keyword Phrase", "Search Volume", "Keyword Sales"])
        st.session_state.cerebro_df = cerebro_df

    competitor_titles = (
        st.session_state.get("competitor_titles")
        or parse_competitor_titles(competitor_text)
    )
    st.session_state.competitor_titles = competitor_titles

    features = [f.strip() for f in features_text.split("\n") if f.strip()]
    synonyms = [s.strip() for s in synonyms_text.split("\n") if s.strip()]
    contexts = [c.strip() for c in contexts_text.split("\n") if c.strip()]

    if is_parent:
        color = ""

    common_args = {
        "lang": lang_code,
        "brand": brand,
        "main_kw": main_kw,
        "size": size,
        "set_size": set_size,
        "material": material,
        "features": features,
        "synonyms": synonyms,
        "contexts": contexts,
        "color": color,
        "no_commas": no_commas,
    }

    titles_data = {}
    for style in ["A", "B", "C"]:
        title = trim_to_fit(style=style, **common_args)
        v = validate_title(title, lang=lang_code, no_commas=no_commas)
        titles_data[style] = {
            "title": title,
            "length": v["length"],
            "valid": v["valid"],
            "issues": v["issues"],
        }

    # Zapis do session state
    st.session_state.generated = True
    st.session_state.titles_data = titles_data
    st.session_state.lang_code = lang_code
    st.session_state.cerebro_df = cerebro_df
    st.session_state.competitor_titles = competitor_titles
    st.session_state.backend_text = backend_text
    st.session_state.no_commas = no_commas
    st.session_state.product_params = {
        "Język": LANGUAGES[lang_code]["name"],
        "Marka": brand,
        "Główna fraza": main_kw,
        "Wymiar": size or "(brak)",
        "Set": set_size or "(brak)",
        "Materiał": material or "(brak)",
        "Cechy": ", ".join(features) or "(brak)",
        "Synonimy": ", ".join(synonyms) or "(brak)",
        "Konteksty": ", ".join(contexts) or "(brak)",
        "Kolor": color if color else "(parent / brak)",
        "Bez przecinków": "Tak" if no_commas else "Nie",
        "Parent-tytuł": "Tak" if is_parent else "Nie",
    }
    st.session_state.translated_data = {}  # reset tłumaczeń

# === Wyświetlanie wyników ===
if st.session_state.generated:
    st.success(f"✅ Wygenerowano tytuły dla: **{LANGUAGES[st.session_state.lang_code]['name']}**")

    for style, label in [
        ("A", "🅰️ A – Klasyczny zlepek (max keywords)"),
        ("B", "🅱️ B – Hybryda (REKOMENDOWANE)"),
        ("C", "🅲 C – Zdaniowa naturalna"),
    ]:
        data = st.session_state.titles_data[style]
        with st.expander(f"{label} ({data['length']} zn.)", expanded=(style == "B")):
            st.code(data["title"], language=None)
            if data["valid"]:
                st.success("✅ Spełnia reguły Amazon")
            else:
                st.error(f"❌ Problemy: {'; '.join(data['issues'])}")
            st.caption(f"Długość: **{data['length']}** / {TITLE_LIMIT} znaków")

    # === Sekcja tłumaczenia ===
    st.markdown("---")
    st.subheader("🌐 Przetłumacz na inny język")
    st.caption(
        "Słownikowe tłumaczenie strukturalne (bez analizy keywordów docelowego rynku). "
        "Działa najlepiej dla branży home & garden. Słowa nieznane są zostawione "
        "w oryginalnym języku - poprawisz je ręcznie albo dodaj do słownika."
    )

    available_targets = [c for c in LANGUAGES.keys() if c != st.session_state.lang_code]
    target_langs = st.multiselect(
        "Wybierz języki docelowe",
        available_targets,
        format_func=lambda x: LANGUAGES[x]["name"],
    )

    col_t1, col_t2 = st.columns([1, 3])
    with col_t1:
        translate_btn = st.button("🔄 Przetłumacz", type="secondary", use_container_width=True)

    if translate_btn and target_langs:
        translated_data = {}
        for target in target_langs:
            translated_titles = {}
            for style in ["A", "B", "C"]:
                src_title = st.session_state.titles_data[style]["title"]
                t_title, unknown = translate_title(
                    src_title, st.session_state.lang_code, target
                )
                translated_titles[style] = {"title": t_title, "unknown": unknown}
            translated_data[target] = {"titles": translated_titles}
        st.session_state.translated_data = translated_data

    if st.session_state.translated_data:
        for target, data in st.session_state.translated_data.items():
            st.markdown(f"### {LANGUAGES[target]['name']}")
            for style in ["A", "B", "C"]:
                info = data["titles"][style]
                with st.expander(f"Opcja {style} ({len(info['title'])} zn.)", expanded=(style == "B")):
                    st.code(info["title"], language=None)
                    if info["unknown"]:
                        st.warning(
                            f"⚠️ Słowa nieznane słownikowi (zostawione w oryginale): "
                            f"`{', '.join(info['unknown'])}`"
                        )
                    else:
                        st.success("✅ Wszystkie słowa przetłumaczone")

    # === Pobieranie Excel ===
    st.markdown("---")
    st.subheader("📥 Pobierz analizę")

    excel_bytes = build_excel(
        titles=st.session_state.titles_data,
        cerebro_df=st.session_state.cerebro_df,
        competitor_titles=st.session_state.competitor_titles,
        backend_terms=st.session_state.backend_text,
        product_params=st.session_state.product_params,
        translated=st.session_state.translated_data if st.session_state.translated_data else None,
    )

    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", st.session_state.product_params["Główna fraza"])[:30] or "tytul"
    st.download_button(
        label="⬇️ Pobierz plik Excel",
        data=excel_bytes,
        file_name=f"Amazon_{st.session_state.lang_code}_tytul_{safe_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # === Dodatkowa analiza ===
    if st.session_state.competitor_titles:
        with st.expander("🔍 Słowa najczęściej używane przez konkurencję", expanded=False):
            kws = extract_keywords_from_titles(
                st.session_state.competitor_titles,
                LANGUAGES[st.session_state.lang_code]["exempt_words"],
            )
            df_kw = pd.DataFrame(kws, columns=["Słowo", "Liczba wystąpień"])
            st.dataframe(df_kw, use_container_width=True, hide_index=True)

    if not st.session_state.cerebro_df.empty:
        with st.expander("📊 Top 20 keywordów z Cerebro", expanded=False):
            display = st.session_state.cerebro_df.head(20)[
                ["Keyword Phrase", "Search Volume", "Keyword Sales"]
            ]
            st.dataframe(display, use_container_width=True, hide_index=True)

else:
    st.info(
        "👈 Wybierz język, wypełnij markę i główną frazę (reszta opcjonalna), "
        "a następnie kliknij **Generuj tytuły**."
    )
